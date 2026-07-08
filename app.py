import os
import json
import secrets
import smtplib
import imaplib
import email
from functools import wraps
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import decode_header
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Flask, render_template, request, jsonify, send_file, session
from werkzeug.security import generate_password_hash, check_password_hash
import anthropic

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))

CONFIG_FILE = "data/config.json"
FORNECEDORES_FILE = "data/fornecedores.json"
COTACOES_FILE = "data/cotacoes.json"
SETORES_FILE = "data/setores.json"
ADMIN_FILE = "data/admin.json"

SETORES_PADRAO = [
    {"id": "ti", "nome": "TI", "cor": "#185FA5", "senha_hash": generate_password_hash("ti123")},
    {"id": "almoxarifado", "nome": "Almoxarifado", "cor": "#3B6D11", "senha_hash": generate_password_hash("almox123")},
    {"id": "manutencao", "nome": "Manutenção", "cor": "#854F0B", "senha_hash": generate_password_hash("manut123")},
    {"id": "escritorio", "nome": "Escritório", "cor": "#6B21A8", "senha_hash": generate_password_hash("escrit123")},
    {"id": "limpeza", "nome": "Limpeza", "cor": "#0F6E56", "senha_hash": generate_password_hash("limpeza123")},
]

ADMIN_PADRAO = {"usuario": "admin", "senha_hash": generate_password_hash("admin123")}

# ── Helpers de persistência ──────────────────────────────────────────────────

def load_json(path, default):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return default

def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_admin():
    return load_json(ADMIN_FILE, ADMIN_PADRAO)

def get_setores():
    return load_json(SETORES_FILE, SETORES_PADRAO)

def setor_publico(s):
    """Remove o hash de senha antes de mandar pro frontend"""
    return {k: v for k, v in s.items() if k != "senha_hash"}

# ── Decorators de proteção de rotas ──────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "tipo" not in session:
            return jsonify({"ok": False, "erro": "Não autenticado"}), 401
        return f(*args, **kwargs)
    return wrapper

def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if session.get("tipo") != "admin":
            return jsonify({"ok": False, "erro": "Acesso restrito ao administrador"}), 403
        return f(*args, **kwargs)
    return wrapper

# ── Autenticação ──────────────────────────────────────────────────────────

@app.route("/api/login", methods=["POST"])
def login():
    data = request.json
    usuario = data.get("usuario", "").strip()
    senha = data.get("senha", "")

    admin = get_admin()
    if usuario == admin["usuario"] and check_password_hash(admin["senha_hash"], senha):
        session["tipo"] = "admin"
        session["usuario"] = usuario
        session.pop("setor_id", None)
        return jsonify({"ok": True, "tipo": "admin", "nome": "Administrador"})

    for s in get_setores():
        if usuario.lower() == s["nome"].lower() and check_password_hash(s.get("senha_hash", ""), senha):
            session["tipo"] = "setor"
            session["setor_id"] = s["id"]
            session["usuario"] = s["nome"]
            return jsonify({"ok": True, "tipo": "setor", "nome": s["nome"], "setor_id": s["id"], "cor": s.get("cor")})

    return jsonify({"ok": False, "erro": "Usuário ou senha incorretos"}), 401

@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})

@app.route("/api/sessao", methods=["GET"])
def sessao():
    if "tipo" not in session:
        return jsonify({"autenticado": False})
    return jsonify({
        "autenticado": True,
        "tipo": session.get("tipo"),
        "usuario": session.get("usuario"),
        "setor_id": session.get("setor_id"),
    })

@app.route("/api/admin/alterar-senha", methods=["POST"])
@admin_required
def alterar_senha_admin():
    nova = request.json.get("senha", "")
    if len(nova) < 4:
        return jsonify({"ok": False, "erro": "Senha muito curta"}), 400
    admin = get_admin()
    admin["senha_hash"] = generate_password_hash(nova)
    save_json(ADMIN_FILE, admin)
    return jsonify({"ok": True})

@app.route("/api/setores/<string:sid>/senha", methods=["POST"])
@admin_required
def alterar_senha_setor(sid):
    nova = request.json.get("senha", "")
    if len(nova) < 4:
        return jsonify({"ok": False, "erro": "Senha muito curta"}), 400
    lista = get_setores()
    for s in lista:
        if s["id"] == sid:
            s["senha_hash"] = generate_password_hash(nova)
    save_json(SETORES_FILE, lista)
    return jsonify({"ok": True})

# ── Rotas principais ─────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/config", methods=["GET", "POST"])
@admin_required
def config():
    if request.method == "POST":
        save_json(CONFIG_FILE, request.json)
        return jsonify({"ok": True})
    return jsonify(load_json(CONFIG_FILE, {}))

@app.route("/api/setores", methods=["GET", "POST"])
@login_required
def setores():
    if request.method == "POST":
        if session.get("tipo") != "admin":
            return jsonify({"ok": False, "erro": "Acesso restrito ao administrador"}), 403
        lista = get_setores()
        novo = request.json
        senha = novo.pop("senha", "setor123")
        novo["id"] = novo["nome"].lower().replace(" ", "_") + f"_{len(lista)}"
        novo["senha_hash"] = generate_password_hash(senha)
        lista.append(novo)
        save_json(SETORES_FILE, lista)
        return jsonify({"ok": True})
    # GET: qualquer usuário logado vê a lista (sem o hash de senha)
    lista = get_setores()
    if session.get("tipo") == "setor":
        lista = [s for s in lista if s["id"] == session.get("setor_id")]
    return jsonify([setor_publico(s) for s in lista])

@app.route("/api/setores/<string:sid>", methods=["DELETE"])
@admin_required
def del_setor(sid):
    lista = get_setores()
    lista = [s for s in lista if s["id"] != sid]
    save_json(SETORES_FILE, lista)
    # remover setor dos fornecedores
    fornecs = load_json(FORNECEDORES_FILE, [])
    for f in fornecs:
        if f.get("setor") == sid:
            f["setor"] = ""
    save_json(FORNECEDORES_FILE, fornecs)
    return jsonify({"ok": True})

@app.route("/api/fornecedores", methods=["GET", "POST"])
@login_required
def fornecedores():
    if request.method == "POST":
        if session.get("tipo") != "admin":
            return jsonify({"ok": False, "erro": "Acesso restrito ao administrador"}), 403
        lista = load_json(FORNECEDORES_FILE, [])
        lista.append(request.json)
        save_json(FORNECEDORES_FILE, lista)
        return jsonify({"ok": True})
    lista = load_json(FORNECEDORES_FILE, [])
    if session.get("tipo") == "setor":
        lista = [f for f in lista if f.get("setor") == session.get("setor_id")]
    return jsonify(lista)

@app.route("/api/fornecedores/<int:idx>", methods=["DELETE"])
@admin_required
def del_fornecedor(idx):
    lista = load_json(FORNECEDORES_FILE, [])
    if 0 <= idx < len(lista):
        lista.pop(idx)
        save_json(FORNECEDORES_FILE, lista)
    return jsonify({"ok": True})

# ── Disparo de e-mails ───────────────────────────────────────────────────────

@app.route("/api/disparar", methods=["POST"])
@login_required
def disparar():
    cfg = load_json(CONFIG_FILE, {})
    todos_fornecs = load_json(FORNECEDORES_FILE, [])
    data = request.json

    # Se logado como setor, força o filtro para o próprio setor (não confia no que o frontend mandar)
    if session.get("tipo") == "setor":
        setor_filtro = session.get("setor_id")
    else:
        setor_filtro = data.get("setor", "")  # admin pode escolher ou deixar vazio = todos

    fornecs = [f for f in todos_fornecs if not setor_filtro or f.get("setor") == setor_filtro]
    assunto_tpl = data.get("assunto", "")
    corpo_tpl = data.get("corpo", "")
    num_cot = data.get("num_cot", "COT-001")
    prazo = data.get("prazo", "5 dias úteis")
    itens = data.get("itens", [])

    tabela_itens = "\n".join(
        f"  {i+1}. {it['produto']} — Qtd: {it['qtd']} {it['unidade']}"
        for i, it in enumerate(itens)
    )

    resultados = []
    erros = []

    try:
        smtp = smtplib.SMTP(cfg["smtp_host"], int(cfg["smtp_port"]))
        smtp.ehlo()
        smtp.starttls()
        smtp.login(cfg["email_user"], cfg["email_pass"])

        for f in fornecs:
            assunto = (assunto_tpl
                .replace("{num_cot}", num_cot)
                .replace("{empresa}", f["nome"]))
            corpo = (corpo_tpl
                .replace("{contato}", f.get("contato", ""))
                .replace("{num_cot}", num_cot)
                .replace("{prazo}", prazo)
                .replace("{remetente}", cfg["email_user"].split("@")[0])
                .replace("{email_user}", cfg["email_user"])
                .replace("{tabela_itens}", tabela_itens)
                .replace("{empresa}", f["nome"]))

            msg = MIMEMultipart()
            msg["From"] = cfg["email_user"]
            msg["To"] = f["email"]
            msg["Subject"] = assunto
            msg.attach(MIMEText(corpo, "plain", "utf-8"))

            try:
                smtp.sendmail(cfg["email_user"], f["email"], msg.as_string())
                resultados.append({"nome": f["nome"], "email": f["email"], "status": "enviado"})
            except Exception as e:
                erros.append({"nome": f["nome"], "email": f["email"], "status": "erro", "msg": str(e)})

        smtp.quit()

        # Registrar cotação no arquivo
        cotacoes = load_json(COTACOES_FILE, [])
        cotacoes.append({
            "num_cot": num_cot,
            "data_disparo": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "itens": itens,
            "fornecedores": [f["nome"] for f in fornecs],
            "setor": setor_filtro or "",
            "criado_por": session.get("usuario", ""),
            "respostas": []
        })
        save_json(COTACOES_FILE, cotacoes)

    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500

    return jsonify({"ok": True, "enviados": resultados, "erros": erros})

# ── Leitura de respostas via IMAP ────────────────────────────────────────────

@app.route("/api/ler-respostas", methods=["POST"])
@login_required
def ler_respostas():
    cfg = load_json(CONFIG_FILE, {})
    num_cot = request.json.get("num_cot", "")

    try:
        mail = imaplib.IMAP4_SSL(cfg["imap_host"], int(cfg.get("imap_port", 993)))
        mail.login(cfg["email_user"], cfg["email_pass"])
        mail.select("INBOX")

        _, ids = mail.search(None, f'SUBJECT "{num_cot}"')
        email_ids = ids[0].split()

        respostas = []
        for eid in email_ids[-20:]:  # últimas 20
            _, msg_data = mail.fetch(eid, "(RFC822)")
            raw = msg_data[0][1]
            msg = email.message_from_bytes(raw)

            assunto_raw, enc = decode_header(msg["Subject"])[0]
            assunto = assunto_raw.decode(enc or "utf-8") if isinstance(assunto_raw, bytes) else assunto_raw

            remetente = msg.get("From", "")
            data_msg = msg.get("Date", "")

            corpo = ""
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type() == "text/plain":
                        payload = part.get_payload(decode=True)
                        corpo = payload.decode(part.get_content_charset() or "utf-8", errors="replace")
                        break
            else:
                payload = msg.get_payload(decode=True)
                corpo = payload.decode(msg.get_content_charset() or "utf-8", errors="replace")

            respostas.append({
                "de": remetente,
                "assunto": assunto,
                "data": data_msg,
                "corpo": corpo[:3000]
            })

        mail.logout()
        return jsonify({"ok": True, "respostas": respostas})

    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500

# ── Análise com IA ───────────────────────────────────────────────────────────

@app.route("/api/analisar", methods=["POST"])
@login_required
def analisar():
    data = request.json
    itens = data.get("itens", [])
    respostas = data.get("respostas", [])
    cotacoes_manuais = data.get("cotacoes_manuais", [])

    itens_txt = "\n".join(f"- {it['produto']}, qtd {it['qtd']} {it['unidade']}" for it in itens)

    if respostas:
        respostas_txt = "\n\n".join(
            f"Fornecedor: {r['de']}\n{r['corpo'][:800]}" for r in respostas
        )
        prompt = f"""Você é um analista de compras. Analise as respostas de cotação abaixo e extraia os preços, prazos e condições de cada fornecedor.

Itens cotados:
{itens_txt}

Respostas recebidas:
{respostas_txt}

Por favor:
1. Monte uma tabela comparativa com preço unitário e total por fornecedor
2. Indique o melhor preço para cada item
3. Dê uma recomendação final de qual fornecedor escolher, considerando preço, prazo e condições
4. Calcule a economia potencial

Responda em português, de forma objetiva e estruturada."""
    else:
        cotacoes_txt = "\n".join(
            f"{c['fornecedor']}: " + ", ".join(f"{k}=R${v}" for k, v in c['precos'].items())
            for c in cotacoes_manuais
        )
        prompt = f"""Analise estas cotações e dê uma recomendação de compra:

Itens: {itens_txt}
Cotações: {cotacoes_txt}

Recomende o melhor fornecedor com justificativa de custo-benefício. Máximo 200 palavras."""

    try:
        client = anthropic.Anthropic()
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1500,
            messages=[{"role": "user", "content": prompt}]
        )
        return jsonify({"ok": True, "analise": response.content[0].text})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500

# ── Exportar Excel ───────────────────────────────────────────────────────────

@app.route("/api/exportar-excel", methods=["POST"])
@login_required
def exportar_excel():
    data = request.json
    num_cot = data.get("num_cot", "COT-001")
    itens = data.get("itens", [])
    cotacoes = data.get("cotacoes", [])
    analise = data.get("analise", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotação"

    # Estilos
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    best_fill = PatternFill("solid", fgColor="D4EDDA")
    title_font = Font(bold=True, size=14, color="1E3A5F")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Título
    ws.merge_cells("A1:G1")
    ws["A1"] = f"COTAÇÃO {num_cot} — {datetime.now().strftime('%d/%m/%Y')}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])

    # Cabeçalho de itens
    ws.append(["#", "Produto", "Qtd", "Unidade"])
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    for i, it in enumerate(itens, 1):
        ws.append([i, it["produto"], it["qtd"], it["unidade"]])
        for cell in ws[ws.max_row]:
            cell.border = border

    ws.append([])
    ws.append([])

    # Tabela comparativa
    fornecs = [c["fornecedor"] for c in cotacoes]
    header_row = ["Item"] + fornecs
    ws.append(header_row)
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for it in itens:
        row = [it["produto"]]
        precos = []
        for c in cotacoes:
            preco = c["precos"].get(it["produto"], 0)
            precos.append(preco)
            row.append(f"R$ {preco:,.2f}" if preco else "—")
        ws.append(row)
        # destacar menor preço
        min_p = min((p for p in precos if p), default=0)
        row_num = ws.max_row
        for col_idx, preco in enumerate(precos, 2):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = border
            if preco == min_p and preco > 0:
                cell.fill = best_fill
                cell.font = Font(bold=True, color="155724")
        ws.cell(row=row_num, column=1).border = border

    # Linha de totais
    ws.append(["TOTAL"] + [
        f"R$ {sum(c['precos'].get(it['produto'], 0) for it in itens):,.2f}"
        for c in cotacoes
    ])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.border = border

    ws.append([])
    ws.append(["PRAZO DE ENTREGA"] + [c.get("prazo", "—") for c in cotacoes])
    ws.append(["CONDIÇÕES DE PAGAMENTO"] + [c.get("pagamento", "—") for c in cotacoes])

    # Aba de análise IA
    if analise:
        ws2 = wb.create_sheet("Análise IA")
        ws2["A1"] = "ANÁLISE E RECOMENDAÇÃO — IA"
        ws2["A1"].font = title_font
        ws2.merge_cells("A1:D1")
        ws2.append([])
        for i, linha in enumerate(analise.split("\n"), 3):
            ws2.cell(row=i, column=1, value=linha)
        ws2.column_dimensions["A"].width = 90

    # Ajuste de colunas
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    path = f"data/cotacao_{num_cot.replace('/', '-')}.xlsx"
    wb.save(path)
    return send_file(path, as_attachment=True, download_name=f"cotacao_{num_cot}.xlsx")

@app.route("/api/cotacoes", methods=["GET"])
@login_required
def listar_cotacoes():
    cotacoes = load_json(COTACOES_FILE, [])
    if session.get("tipo") == "setor":
        cotacoes = [c for c in cotacoes if c.get("setor") == session.get("setor_id")]
    return jsonify(list(reversed(cotacoes)))

# ── Teste de conexão SMTP ────────────────────────────────────────────────────

@app.route("/api/testar-conexao", methods=["POST"])
@admin_required
def testar_conexao():
    cfg = request.json
    try:
        smtp = smtplib.SMTP(cfg["smtp_host"], int(cfg["smtp_port"]), timeout=8)
        smtp.ehlo()
        smtp.starttls()
        smtp.login(cfg["email_user"], cfg["email_pass"])
        smtp.quit()
        return jsonify({"ok": True, "msg": "Conexão SMTP bem-sucedida!"})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 400

if __name__ == "__main__":
    os.makedirs("data", exist_ok=True)
    app.run(host="0.0.0.0", port=5000, debug=True)
